import os
import sqlite3
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox
from openpyxl import Workbook

class Application(tk.Frame):
    """CLASSE PARA ABRIR BANCO DE DADOS EM SQLITE COM ANDAMENTO DOS PROCESSOS PESQUISADOS PELA 
    MÓDULO pesquisar_processo_selenium. É POSSÍVEL FAZER PESQUISAS DE TEXTO NO ANDAMENTO DO
    PROCESSO E EXPORTAR PROCESSOS SELECIONADOS"""
    def __init__(self, master=None):
        super().__init__(master)
        self.master = master
        self.pack(fill="both", expand=1)
        self.diretorio = "Precatorios_2013_06.05.2017"
        self.processos_visualizados = set()
        self.processos_selecionados = set()
        self.create_widgets()

    def create_widgets(self):

        #MENU
        self.menubar = tk.Menu(self)
        menu = tk.Menu(self.menubar, tearoff=0)
        self.menubar.add_cascade(label="Arquivo", menu=menu)
        menu.add_command(label="Abrir banco de dados", command=self.open_db)
        menu.add_command(label="Fechar banco de dados", command=self.close_db)
        try:
            self.master.config(menu=self.menubar)
        except AttributeError:
            # master is a toplevel window (Python 1.4/Tkinter 1.63)
            self.master.tk.call(master, "config", "-menu", self.menubar)

        #CANVAS FOOTER
        self.canvas_footer = tk.Canvas(self)
        self.canvas_footer.pack(side="bottom", fill="x", padx="5", pady="5")
        self.lbl_vara = tk.Label(self.canvas_footer, text="Vara: ", width=30, anchor="w")
        self.lbl_vara.pack(side="left", padx="5", pady="5")
        self.lbl_dias = tk.Label(self.canvas_footer,
                                 text="Última movimentação há: ",
                                 width=30, anchor="w")
        self.lbl_dias.pack(side="left", padx="5", pady="5")
        self.checkbutton_selected_intvar = tk.IntVar()
        self.checkbutton_selected = tk.Checkbutton(self.canvas_footer,
                                                   text="Processo selecionado",
                                                   variable=self.checkbutton_selected_intvar,
                                                   command=self.selecionar_processo,
                                                   takefocus=False)
        self.checkbutton_selected.pack(side="left", padx="5", pady="5")
        self.master.bind_all('<space>', self.toggle_checkbutton)
        self.btn_exportar = tk.Button(self.canvas_footer,
                                      command=self.exportar,
                                      takefocus=False)
        self.btn_exportar_change()
        self.btn_exportar.pack(side="left", padx="5", pady="5")

        #CANVAS HEADER ESQUERDA
        self.canvas_head = tk.Canvas(self)
        self.canvas_head.pack(side="top", fill="x", padx="5", pady="5")
        self.lbl_db = tk.Label(self.canvas_head, text="", width=30)
        self.lbl_db.pack(side="left", padx="5", pady="5")
        self.separator1 = ttk.Separator(self.canvas_head, orient="vertical")
        self.separator1.pack(side="left", fill="y")
        self.lbl_qnt_processos = tk.Label(self.canvas_head, text="", width=30)
        self.lbl_qnt_processos.pack(side="left", padx="5", pady="5")
        self.separator2 = ttk.Separator(self.canvas_head, orient="vertical")
        self.separator2.pack(side="left", fill="y")
        self.lbl_pesquisa = tk.Label(self.canvas_head, text="", width=70)
        self.lbl_pesquisa.pack(side="left", padx="5", pady="5")

        #CANVAS HEADER DIREITA
        self.btn_limpar_pesquisa = tk.Button(self.canvas_head,
                                             text="Limpar",
                                             takefocus=False,
                                             command=self.limpar_pesquisa)
        self.btn_limpar_pesquisa.pack(side="right", padx="5", pady="5")
        self.btn_pesquisar = tk.Button(self.canvas_head,
                                       text="Pesquisar",
                                       takefocus=False,
                                       command=self.pesquisar)
        self.btn_pesquisar.pack(side="right", padx="5", pady="5")
        self.entry_pesquisa_stringvar = tk.StringVar()
        self.entry_pesquisa = ttk.Combobox(self.canvas_head,
                                       takefocus=False,
                                       textvariable=self.entry_pesquisa_stringvar)
        self.entry_pesquisa.pack(side="right", padx="5", pady="5")
        self.entry_pesquisa.bind('<Return>', lambda event: self.pesquisar())
        self.separator3 = ttk.Separator(self.canvas_head, orient="vertical")
        self.separator3.pack(side="right", fill="y")

        #CANVAS LISTBOX
        self.canvas_listbox = tk.Canvas(self)
        self.canvas_listbox.pack(side="left",
                                 anchor="w",
                                 fill="y",
                                 expand=False,
                                 padx="10",
                                 pady="10")
        self.listbox_processos = tk.Listbox(self.canvas_listbox, width=30)
        self.listbox_processos.pack(side="left", fill="y", expand=False)
        self.scrollbar_lista_processos = tk.Scrollbar(self.canvas_listbox,
                                                      takefocus=False,
                                                      orient="vertical")
        self.scrollbar_lista_processos.config(command=self.listbox_processos.yview)
        self.scrollbar_lista_processos.pack(side="right", fill="y")
        self.listbox_processos.config(yscrollcommand=self.scrollbar_lista_processos.set)
        self.listbox_processos.bind("<<ListboxSelect>>", self.selected_listbox)

        #CANVAS TEXT
        self.canvas_text = tk.Canvas(self)
        self.canvas_text.pack(side="left",
                              fill="both",
                              expand=True,
                              padx="10",
                              pady="10")
        self.text_andamentos = tk.Text(self.canvas_text)
        self.text_andamentos.pack(side="left", fill="both", expand=True)
        self.scrollbar_andamentos = tk.Scrollbar(self.canvas_text, orient="vertical")
        self.scrollbar_andamentos.config(command=self.text_andamentos.yview)
        self.scrollbar_andamentos.pack(side="right", fill="y")
        self.text_andamentos.config(yscrollcommand=self.scrollbar_andamentos.set,
                                    state="disabled")

    def open_db(self):
        #ABRE UM NOVO BANCO DE DADOS SQLITE
        try:
            self.close_db()
        except:
            pass
        self.processos_visualizados = set()
        self.processos_selecionados = set()
        self.termos_pesquisados = set()
        self.entry_pesquisa["values"] = list(self.termos_pesquisados)
        self.actualpath = os.getcwd()
        self.path = filedialog.askopenfilename(initialdir=self.actualpath,
                                               title='Escolha uma pasta',
                                               defaultextension=".db",
                                               filetypes=[('banco de dados','.db')])
        self.conn = sqlite3.connect(self.path)
        self.cur = self.conn.cursor()
        self.cur.execute('''SELECT numero_processo FROM processos ORDER BY dias_parados DESC''')
        self.processos = self.cur.fetchall()
        for processo, in self.processos:
            self.listbox_processos.insert(tk.END, str(processo))
        #self.listbox_processos.focus_set()
        #self.listbox_processos.select_set(0)  # This only sets focus on the first item.
        #self.listbox_processos.event_generate("<<ListboxSelect>>")
        self.lbl_db.config(text=os.path.basename(self.path))
        self.lbl_qnt_processos.config(text="Foram encontrados " + str(len(self.processos)) + " processos")

    def close_db(self):
        #FECHA O BANCO DE DADOS SQLITE ATUAL
        self.conn.commit()
        self.conn.close()
        self.checkbutton_selected_intvar.set(0)
        self.processos_visualizados = set()
        self.processos_selecionados = set()
        self.termos_pesquisados = set()
        self.entry_pesquisa["values"] = list(self.termos_pesquisados)
        self.btn_exportar_change()
        self.listbox_processos.delete(0, tk.END)
        self.text_andamentos.config(state="normal")
        self.text_andamentos.delete(1.0, tk.END)
        self.text_andamentos.config(state="disabled")
        self.lbl_vara.config(text="Vara: ")
        self.lbl_dias.config(text="Última movimentação há: ")
        self.lbl_db.config(text="")
        self.lbl_pesquisa.config(text="")
        self.lbl_qnt_processos.config(text="")

    def selected_listbox(self, event):
        #ABRE A MOVIMENTAÇÃO DO PROCESSO SELECIONADO
        self.widget = event.widget
        self.selection = self.widget.curselection()
        self.value = self.widget.get(self.selection[0])
        self.fill_text(self.value)
        self.lbl_vara.config(text='Vara: '+ self.vara)
        self.lbl_dias.config(text="Última movimentação há: " + str(self.dias_parados) + " dias")
        self.text_andamentos.config(state="normal")
        self.text_andamentos.delete(1.0,tk.END)
        for numero_processo, data, andamento in self.andamentos:
            self.text_andamentos.insert(tk.END, data+"\n")
            self.text_andamentos.insert(tk.END, andamento.replace("ÀS","\nÀS")+"\n")
            self.text_andamentos.insert(tk.END, "\n"+"#"*50+"\n\n")
        self.highlight_pesquisa()
        self.text_andamentos.config(state="disabled")
        #VERIFICA SE O PROCESSO ESTÁ SELECIONADO E MARCA O CHECKBOX
        if self.listbox_processos.itemcget(self.widget.curselection(), 'bg') == 'yellow':
            self.checkbutton_selected_intvar.set(1)
        else:
            self.checkbutton_selected_intvar.set(0)
        #MARCA OS PROCESSOS VISTOS
        self.listbox_processos.itemconfig(self.selection[0], fg='blue')
        self.processos_visualizados.add(self.value)

    def fill_text(self, processo):
        #PESQUISA O PROCESSO E OS ANDAMENTOS DO PROCESSO SELECIONADO
        self.cur.execute('''SELECT * FROM andamentos WHERE numero_processo=?''', (processo,))
        self.andamentos = self.cur.fetchall()
        self.cur.execute('''SELECT vara, dias_parados FROM processos WHERE numero_processo=?''', (processo,))
        self.vara, self.dias_parados = self.cur.fetchone()

    def pesquisar(self):
        #REALIZA UMA PESQUISA NO BANCO DE DADOS
        self.btn_exportar_change()
        self.str_pesquisa = self.entry_pesquisa_stringvar.get()
        self.cur.execute("""SELECT DISTINCT a.numero_processo 
                        FROM andamentos AS a JOIN processos AS p 
                        ON p.numero_processo = a.numero_processo
                        WHERE a.andamento LIKE ? ORDER BY p.dias_parados DESC""",
                         ('%'+self.str_pesquisa+'%',))
        self.resultado_pesquisa = self.cur.fetchall()
        #print(self.resultado_pesquisa)
        self.listbox_processos.delete(0, tk.END)
        for processo, in self.resultado_pesquisa:
            self.listbox_processos.insert(tk.END, str(processo))
        #self.listbox_processos.select_set(0)  # This only sets focus on the first item.
        #self.listbox_processos.event_generate("<<ListboxSelect>>")
        #self.listbox_processos.focus_set()
        self.lbl_pesquisa.config(text="Você está pesquisando pelo termo '" + self.str_pesquisa + "'")
        self.lbl_qnt_processos.config(text="Foram encontrados " + str(len(self.resultado_pesquisa)) + " processos")
        self.marcar_processos_selecionados_e_vistos()
        #ADICIONA O TERMO PESQUISADO AO COMBOBOX DA PESQUISA
        self.termos_pesquisados.add(self.str_pesquisa)
        self.entry_pesquisa["values"] = list(self.termos_pesquisados)

    def highlight_pesquisa(self):
        #SINALIZA DE AMARELO O TEXTO PESQUISADO NOS ANDAMENTOS
        self.str_pesquisa = self.entry_pesquisa_stringvar.get()
        if self.str_pesquisa != "":
            self.start_pos = self.text_andamentos.search(self.str_pesquisa, '1.0', stopindex=tk.END, nocase=True)
            #print(self.start_pos)
            if self.start_pos:
                self.end_pos = '{}+{}c'.format(self.start_pos, len(self.str_pesquisa))
                #print(self.end_pos)
                self.text_andamentos.config(state="normal")
                self.text_andamentos.tag_add('highlight', str(self.start_pos), str(self.end_pos))
                self.text_andamentos.tag_config('highlight', background='yellow')
                self.text_andamentos.config(state="disabled")

    def limpar_pesquisa(self):
        #LIMPA A PESQUISA ATUAL E RETORNA AO BANCO DE DADOS ABERTO
        self.btn_exportar_change()
        self.listbox_processos.delete(0, tk.END)
        self.text_andamentos.delete(1.0, tk.END)
        self.entry_pesquisa.delete(0, tk.END)
        self.cur.execute('''SELECT numero_processo FROM processos ORDER BY dias_parados DESC''')
        self.processos = self.cur.fetchall()
        for processo, in self.processos:
            self.listbox_processos.insert(tk.END, str(processo))
        #self.listbox_processos.select_set(0)  # This only sets focus on the first item.
        #self.listbox_processos.event_generate("<<ListboxSelect>>")
        #self.listbox_processos.focus_set()
        self.lbl_qnt_processos.config(text="Foram encontrados " + str(len(self.processos)) + " processos")
        self.lbl_pesquisa.config(text="")
        self.marcar_processos_selecionados_e_vistos()

    def marcar_processos_selecionados_e_vistos(self):
        #MARCA OS PROCESSOS JÁ VISTOS E SELECIONADOS APÓS UMA PESQUISA
        for i, listbox_entry in enumerate(self.listbox_processos.get(0, tk.END)):
            if listbox_entry in self.processos_selecionados:
                self.listbox_processos.itemconfig(i, bg='yellow')
            if listbox_entry in self.processos_visualizados:
                self.listbox_processos.itemconfig(i, fg='blue') 

    def toggle_checkbutton(self, event):
        #ALTERA O CHECKBUTTON DE PROCESSO SELECIONADO
        self.checkbutton_selected.toggle()
        self.selecionar_processo()

    def selecionar_processo(self):
        #PINTA O ITEM NO LISTBOX DE PROCESSO SEMPRE QUE O CHECKBOX É MARCADO
        self.lstbox_item_idx = self.listbox_processos.curselection()
        if self.checkbutton_selected_intvar.get() == 1:
            self.listbox_processos.itemconfig(self.lstbox_item_idx, bg='yellow')
            self.processos_selecionados.add(self.listbox_processos.get(self.lstbox_item_idx))
            self.btn_exportar_change()
        else:
            self.listbox_processos.itemconfig(self.lstbox_item_idx, bg='white')
            self.processos_selecionados.remove(self.listbox_processos.get(self.lstbox_item_idx))
            self.btn_exportar_change()

    def btn_exportar_change(self):
        #ALTERA O TEXTO NO BOTÃO EXPORTAR
        self.btn_exportar_qnt = len(self.processos_selecionados)
        self.btn_exportar_str = "Exportar {} processos selecionados".format(self.btn_exportar_qnt)
        self.btn_exportar.configure(text=self.btn_exportar_str)

    def exportar(self):
        #EXPORTA OS PROCESSOS SELECIONADOS
        self.wb = Workbook()
        self.wb_filename = filedialog.asksaveasfilename(initialdir=self.actualpath,
            title='Escolha o nome do arquivo', defaultextension=".xlsx", filetypes=[('Excel','.xlsx')])
        self.ws = self.wb.active
        self.ws.title = "Processos"
        self.ws.cell(row=1, column=1).value = "PROCESSOS"
        contador = 1
        for processo in self.processos_selecionados:
            self.ws.cell(row=contador+1, column=1).value = processo
            contador += 1
        self.wb.save(self.wb_filename)
        messagebox.showinfo("Exportar processos", "Processos selecionados exportados com sucesso!")

root = tk.Tk()
root.state('zoomed')
root.title("Relatórios Processuais")
app = Application(master=root)
app.mainloop()
