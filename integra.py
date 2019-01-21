import os
import sqlite3
import locale
import time
import re
from datetime import datetime
from bs4 import BeautifulSoup
from shutil import copyfile
from socket import create_connection
from traceback import print_exc
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from openpyxl import load_workbook
from PIL import Image, ImageFilter
import pytesseract
from solve_captcha import solve_captcha_jurisconsult, solve_captcha_pje

try:
    locale.setlocale(locale.LC_ALL, 'pt_BR')
except:
    locale.setlocale(locale.LC_ALL, 'Portuguese_Brazil')

class DESC_PG():
    '''Classe para marcar os clientes que já pagaram os honorários iniciais do processo de descompressão '''
    def __init__(self, lista_processos='', tipo_de_processo='', diretorio=''):
        self.lista_processos = lista_processos
        self.tipo_de_processo = tipo_de_processo
        self.diretorio = diretorio
        self.driver = webdriver.Chrome('chromedriver.exe')
        self.driver.get('http://jurisconsult.tjma.jus.br')
        self.driver.implicitly_wait(3)
        if not os.path.exists("data/" + self.diretorio):
            os.makedirs("data/" + self.diretorio)
            os.makedirs("data/" + self.diretorio + "/html")
            os.makedirs("data/" + self.diretorio + "/error")
        self.open_db()
        try:
            self.contador = 1
            self.message = ""
            for processo, in self.lista_processos:                                
                if (str(processo.value).startswith("08")) and (len(str(processo.value))>10):
                    continue
                while True:
                    if self.tipo_de_processo == "primeiro_grau":
                        self.inserido = self.inserir_numero_primeiro_grau(processo.value)
                    elif self.tipo_de_processo == "precatorio":
                        self.inserido = self.inserir_numero_precatorio(processo.value)
                    else:
                        raise Exception("O tipo_de_processo deve ser 'primeiro_grau' ou 'precatorio'!")
                    if self.message != "Informação, da imagem, não coincide. Favor digite novamente!":
                        break
                if "Consulta realizada com sucesso" in self.message:
                    self.html = self.driver.page_source
                    with open("data/" + self.diretorio+'/html/' + str(processo.value) + '.html', 'w') as escrita:
                        escrita.write(self.html)
                    self.salva_banco(self.html, processo.value, self.message)
                self.contador = self.contador + 1                  
        except NoSuchElementException:
            if self.is_connected() == False:
                print("Sem conexão com a internet!")
            else:
                self.driver.save_screenshot("data/" + self.diretorio+"/error/last_exception.png")
                print_exc()
        except:
            self.driver.save_screenshot("data/" + self.diretorio + "/error/last_exception.png")
            print_exc()
        finally:
            self.conn.commit()
            self.conn.close()
            self.driver.close()
            self.driver.quit()
            pass

    def is_connected(self):
        try:
            create_connection(("www.google.com", 80))
            return True
        except OSError:
            pass
        return False

    def open_db(self):
        self.conn = sqlite3.connect("data/" + self.diretorio + '/' + self.diretorio + '.db')
        self.cur = self.conn.cursor()
        self.cur.execute('''CREATE TABLE IF NOT EXISTS processos
                    (id INTEGER PRIMARY KEY AUTOINCREMENT,
                    numero_processo TEXT UNIQUE, vara TEXT, dias_parados INTEGER, mensagem_download TEXT)''')
        self.cur.execute('''CREATE TABLE IF NOT EXISTS andamentos (numero_processo TEXT, data TEXT, 
                    andamento TEXT, FOREIGN KEY(numero_processo) REFERENCES processos(numero_processo))''')

    def inserir_numero_primeiro_grau(self, numero):
        if self.contador == 1:
            self.driver.find_element_by_link_text('Primeiro Grau').click()
            self.driver.find_element_by_link_text('Consulta Pública').click()
            try:
                self.driver.find_element_by_xpath("//div[@class='ui-dialog-buttonpane ui-widget-content ui-helper-clearfix']/div/button/span[@class='ui-button-text']").click()
            except:
                pass
        elif "Consulta realizada com sucesso" in self.message:
                self.driver.find_element_by_xpath("//a[@class='abrirAbas']/img[@title='abrir/esconder formulario']").click()
        if len(str(numero))>10:
            ####NUMERAÇÃO CNJ
            self.driver.find_element_by_xpath('//select[@id="slcTipo"]/option[@value="2"]').click()
        self.driver.find_element_by_id('txtChave').send_keys(numero)
        self.captcha_element = self.driver.find_element_by_xpath("//img[@src='http://jurisconsult.tjma.jus.br/app/views/elements/captcha.php']")
        self.captcha = self.handle_captcha_jurisconsult(self.captcha_element, numero)
        self.driver.find_element_by_id('txtCaptcha').send_keys(self.captcha)
        self.driver.find_element_by_id('btnConsultar').click()
        try:
            self.message = self.driver.find_element_by_xpath("//div[@class='message']/div/p").text
        except:
            self.message = ""
        print("Processo:", numero, "-", self.message)
        if self.message == "Informação, da imagem, não coincide. Favor digite novamente!":
            copyfile("data/" + self.diretorio + "/captcha_" + str(numero) + ".png", "data/" + self.diretorio + "/error/captcha_" + str(numero) + ".png")
        os.remove("data/" + self.diretorio + "/captcha_" + str(numero) + ".png")

    def inserir_numero_precatorio(self, numero):
        numero = str(numero).strip()
        while len(numero) < 10:
            numero = "0"+numero
        if self.contador == 1:
            self.driver.find_element_by_id('segundoGrau').click()
            self.driver.find_element_by_id('precatorio').click()
        elif "Consulta realizada com sucesso" in self.message:
            self.driver.find_element_by_xpath("//a[@class='abrirAbas']/img[@title='abrir/esconder formulario']").click()
        self.driver.find_element_by_id('txtChave').clear()
        self.driver.find_element_by_id('txtChave').send_keys(numero)
        self.driver.find_element_by_id('btnConsultar').click()
        try:
            self.message = self.driver.find_element_by_xpath("//div[@class='message']/div/p").text
        except:
            self.message = ""
        print("Processo:", numero, "-", self.message)

    def handle_captcha_jurisconsult(self, img, numero):
        self.screenshot_filename = "data/" + self.diretorio + "/screenshot_" + str(numero) + ".png"
        self.captcha_filename = "data/" + self.diretorio + "/captcha_" + str(numero) + ".png"
        self.location = img.location
        self.size = img.size
        self.driver.save_screenshot(self.screenshot_filename)
        self.screenshot = Image.open(self.screenshot_filename)
        self.left = self.location['x']
        self.top = self.location['y']
        self.right = self.location['x'] + self.size['width']
        self.bottom = self.location['y'] + self.size['height']
        self.captcha_img = self.screenshot.crop((self.left, self.top, self.right, self.bottom))
        self.screenshot.close()
        self.captcha_img, solution = solve_captcha_jurisconsult(self.captcha_img)
        self.captcha_img.save(self.captcha_filename)
        os.remove(self.screenshot_filename)
        return solution

    def salva_banco(self, html, processo, message):
        self.page_soup = BeautifulSoup(html, 'html.parser')
        if self.tipo_de_processo == "primeiro_grau":
            self.vara = self.page_soup.find(string="Vara:").find_next("td").text
        elif self.tipo_de_processo == "precatorio":
            self.vara = "Tribunal de Justiça"
        try:
            self.cur.execute("""INSERT INTO processos (numero_processo, vara, mensagem_download) VALUES
                            (?, ?, ?)""", (processo, self.vara, message))
        except sqlite3.IntegrityError:
            print_exc()
            return
        self.primeira_movimentacao = self.page_soup.find(class_="movimentacao")
        self.ultima_data = self.primeira_movimentacao.h1.text.replace("\n", "").replace("\t", "")
        self.qnt_dias = datetime.now() - datetime.strptime(self.ultima_data, "%A, %d de %B de %Y")
        self.qnt_dias = str(self.qnt_dias.days)
        self.cur.execute("""UPDATE processos SET
                    dias_parados = ? WHERE numero_processo = ?""", (self.qnt_dias, processo))
        self.movimentacoes = self.page_soup.find_all(class_="movimentacao")
        for movimentacao in self.movimentacoes:
            self.data = movimentacao.h1.text.replace("\n", "").replace("\t", "")
            movimentacao.h1.decompose()
            self.movimentacao = ' '.join(movimentacao.text.split())
            self.cur.execute("""INSERT INTO andamentos (numero_processo, data, andamento) VALUES 
                            (?, ?, ?)""", (processo, self.data, self.movimentacao))


class PesquisarClientes():
    '''Classe para salvar as fichas de uma lista de clientes do INTEGRA.'''
    def __init__(self, lista_clientes='', diretorio=''):
        self.lista_clientes = lista_clientes
        self.diretorio = diretorio
        self.driver = webdriver.Chrome('chromedriver.exe')
        self.driver.get('https://www.integra.adv.br')
        self.email = os.environ.get('INTEGRA_EMAIL', 'your-email@example.com')
        self.senha = os.environ.get('INTEGRA_PASSWORD', 'your-password')
        self.driver.implicitly_wait(3)
        if not os.path.exists("data/" + self.diretorio):
            os.makedirs("data/" + self.diretorio)
            os.makedirs("data/" + self.diretorio + "/html")
            os.makedirs("data/" + self.diretorio + "/error")
        try:
            self.login()
            self.contador = 1
            self.message = ""
            for cliente, in self.lista_clientes:
                self.pesquisar(cliente)
                self.contador = self.contador + 1
        except NoSuchElementException:
            if self.is_connected() == False:
                print("Sem conexão com a internet!")
            else:
                self.driver.save_screenshot("data/" + self.diretorio + "/error/last_exception.png")
                print_exc()
        except:
            self.driver.save_screenshot("data/" + self.diretorio + "/error/last_exception.png")
            print_exc()
        finally:
            self.logout()
            self.driver.close()
            self.driver.quit()
            pass

    def is_connected(self):
        try:
            create_connection(("www.google.com", 80))
            return True
        except OSError:
            pass
        return False

    def login(self):
        self.driver.find_element_by_id('login_email').send_keys(self.email)
        self.driver.find_element_by_id('login_senha').send_keys(self.senha)
        self.driver.find_element_by_xpath("//button[@type='submit']").click()
        self.driver.find_element_by_xpath("//a[text()='Fechar(X)']").click()

    def logout(self):
        self.driver.find_element_by_xpath("//p[@class='menufechar']").click()

    def pesquisar(self, cliente):
        self.driver.find_element_by_xpath("//p[text()='  Pesquisar Cliente   ']").click()

if __name__ == '__main__':
    # wb = load_workbook(filename='excel/PRECATORIOS a partir de 2012 - 30 Ago 2017.xlsx', read_only=True)
    # #ws = wb[wb.sheetnames[0]]
    # ws = wb['TODOS']
    # range = ws['A1':'A672']
    # IncluirPush = PesquisarProcessoJurisconsult(lista_processos=range,
    #                                         tipo_de_processo="precatorio",
    #                                         diretorio='PRECATORIOS a partir de 2012 - 30 Ago 2017')

    wb = load_workbook(filename='excel/CLIENTES - Pedreiras e Região.xlsx', read_only=True)
    #ws = wb['PJE']
    ws = wb[wb.sheetnames[0]]
    range = ws['A2':'A1320']
    IncluirPush = PesquisarClientes(lista_clientes=range,
                                        diretorio='CLIENTES - Pedreiras e Região')