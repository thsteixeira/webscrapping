import os
import sqlite3
import locale
import time
import re
from datetime import datetime
from bs4 import BeautifulSoup
from shutil import copyfile
from socket import create_connection
from traceback import print_exc
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from openpyxl import load_workbook
from PIL import Image, ImageFilter
import pytesseract
from solve_captcha import solve_captcha_jurisconsult, solve_captcha_pje

try:
    locale.setlocale(locale.LC_ALL, 'pt_BR')
except:
    locale.setlocale(locale.LC_ALL, 'Portuguese_Brazil')

class PesquisarProcessoJurisconsult():
    '''Classe para realizar pesquisa no site Jurisconsult dos processos listados em 'lista_processos',
    podendo serem pesquisados em 'primeiro_grau' ou em 'precatorio'. Os resultados serão 
    salvos na pasta 'diretorio'. '''
    def __init__(self, lista_processos='', tipo_de_processo='', diretorio=''):
        self.lista_processos = lista_processos
        self.tipo_de_processo = tipo_de_processo
        self.diretorio = diretorio
        self.driver = webdriver.Chrome('chromedriver.exe')
        self.driver.get('http://jurisconsult.tjma.jus.br')
        self.driver.implicitly_wait(3)
        if not os.path.exists("data/" + self.diretorio):
            os.makedirs("data/" + self.diretorio)
            os.makedirs("data/" + self.diretorio + "/html")
            os.makedirs("data/" + self.diretorio + "/error")
        self.open_db()
        try:
            self.contador = 1
            self.message = ""
            for processo, in self.lista_processos:                                
                if (str(processo.value).startswith("08")) and (len(str(processo.value))>10):
                    continue
                while True:
                    if self.tipo_de_processo == "primeiro_grau":
                        self.inserido = self.inserir_numero_primeiro_grau(processo.value)
                    elif self.tipo_de_processo == "precatorio":
                        self.inserido = self.inserir_numero_precatorio(processo.value)
                    else:
                        raise Exception("O tipo_de_processo deve ser 'primeiro_grau' ou 'precatorio'!")
                    if self.message != "Informação, da imagem, não coincide. Favor digite novamente!":
                        break
                if "Consulta realizada com sucesso" in self.message:
                    self.html = self.driver.page_source
                    with open("data/" + self.diretorio+'/html/' + str(processo.value) + '.html', 'w') as escrita:
                        escrita.write(self.html)
                    self.salva_banco(self.html, processo.value, self.message)
                self.contador = self.contador + 1                  
        except NoSuchElementException:
            if self.is_connected() == False:
                print("Sem conexão com a internet!")
            else:
                self.driver.save_screenshot("data/" + self.diretorio+"/error/last_exception.png")
                print_exc()
        except:
            self.driver.save_screenshot("data/" + self.diretorio + "/error/last_exception.png")
            print_exc()
        finally:
            self.conn.commit()
            self.conn.close()
            self.driver.close()
            self.driver.quit()
            pass

    def is_connected(self):
        try:
            create_connection(("www.google.com", 80))
            return True
        except OSError:
            pass
        return False

    def open_db(self):
        self.conn = sqlite3.connect("data/" + self.diretorio + '/' + self.diretorio + '.db')
        self.cur = self.conn.cursor()
        self.cur.execute('''CREATE TABLE IF NOT EXISTS processos
                    (id INTEGER PRIMARY KEY AUTOINCREMENT,
                    numero_processo TEXT UNIQUE, vara TEXT, dias_parados INTEGER, mensagem_download TEXT)''')
        self.cur.execute('''CREATE TABLE IF NOT EXISTS andamentos (numero_processo TEXT, data TEXT, 
                    andamento TEXT, FOREIGN KEY(numero_processo) REFERENCES processos(numero_processo))''')

    def inserir_numero_primeiro_grau(self, numero):
        if self.contador == 1:
            self.driver.find_element_by_link_text('Primeiro Grau').click()
            self.driver.find_element_by_link_text('Consulta Pública').click()
            try:
                self.driver.find_element_by_xpath("//div[@class='ui-dialog-buttonpane ui-widget-content ui-helper-clearfix']/div/button/span[@class='ui-button-text']").click()
            except:
                pass
        elif "Consulta realizada com sucesso" in self.message:
                self.driver.find_element_by_xpath("//a[@class='abrirAbas']/img[@title='abrir/esconder formulario']").click()
        if len(str(numero))>10:
            ####NUMERAÇÃO CNJ
            self.driver.find_element_by_xpath('//select[@id="slcTipo"]/option[@value="2"]').click()
        self.driver.find_element_by_id('txtChave').send_keys(numero)
        self.captcha_element = self.driver.find_element_by_xpath("//img[@src='http://jurisconsult.tjma.jus.br/app/views/elements/captcha.php']")
        self.captcha = self.handle_captcha_jurisconsult(self.captcha_element, numero)
        self.driver.find_element_by_id('txtCaptcha').send_keys(self.captcha)
        self.driver.find_element_by_id('btnConsultar').click()
        try:
            self.message = self.driver.find_element_by_xpath("//div[@class='message']/div/p").text
        except:
            self.message = ""
        print("Processo:", numero, "-", self.message)
        if self.message == "Informação, da imagem, não coincide. Favor digite novamente!":
            copyfile("data/" + self.diretorio + "/captcha_" + str(numero) + ".png", "data/" + self.diretorio + "/error/captcha_" + str(numero) + ".png")
        os.remove("data/" + self.diretorio + "/captcha_" + str(numero) + ".png")

    def inserir_numero_precatorio(self, numero):
        numero = str(numero).strip()
        while len(numero) < 10:
            numero = "0"+numero
        if self.contador == 1:
            self.driver.find_element_by_id('segundoGrau').click()
            self.driver.find_element_by_id('precatorio').click()
        elif "Consulta realizada com sucesso" in self.message:
            self.driver.find_element_by_xpath("//a[@class='abrirAbas']/img[@title='abrir/esconder formulario']").click()
        self.driver.find_element_by_id('txtChave').clear()
        self.driver.find_element_by_id('txtChave').send_keys(numero)
        self.driver.find_element_by_id('btnConsultar').click()
        try:
            self.message = self.driver.find_element_by_xpath("//div[@class='message']/div/p").text
        except:
            self.message = ""
        print("Processo:", numero, "-", self.message)

    def handle_captcha_jurisconsult(self, img, numero):
        self.screenshot_filename = "data/" + self.diretorio + "/screenshot_" + str(numero) + ".png"
        self.captcha_filename = "data/" + self.diretorio + "/captcha_" + str(numero) + ".png"
        self.location = img.location
        self.size = img.size
        self.driver.save_screenshot(self.screenshot_filename)
        self.screenshot = Image.open(self.screenshot_filename)
        self.left = self.location['x']
        self.top = self.location['y']
        self.right = self.location['x'] + self.size['width']
        self.bottom = self.location['y'] + self.size['height']
        self.captcha_img = self.screenshot.crop((self.left, self.top, self.right, self.bottom))
        self.screenshot.close()
        self.captcha_img, solution = solve_captcha_jurisconsult(self.captcha_img)
        self.captcha_img.save(self.captcha_filename)
        os.remove(self.screenshot_filename)
        return solution

    def salva_banco(self, html, processo, message):
        self.page_soup = BeautifulSoup(html, 'html.parser')
        if self.tipo_de_processo == "primeiro_grau":
            self.vara = self.page_soup.find(string="Vara:").find_next("td").text
        elif self.tipo_de_processo == "precatorio":
            self.vara = "Tribunal de Justiça"
        try:
            self.cur.execute("""INSERT INTO processos (numero_processo, vara, mensagem_download) VALUES
                            (?, ?, ?)""", (processo, self.vara, message))
        except sqlite3.IntegrityError:
            print_exc()
            return
        self.primeira_movimentacao = self.page_soup.find(class_="movimentacao")
        self.ultima_data = self.primeira_movimentacao.h1.text.replace("\n", "").replace("\t", "")
        self.qnt_dias = datetime.now() - datetime.strptime(self.ultima_data, "%A, %d de %B de %Y")
        self.qnt_dias = str(self.qnt_dias.days)
        self.cur.execute("""UPDATE processos SET
                    dias_parados = ? WHERE numero_processo = ?""", (self.qnt_dias, processo))
        self.movimentacoes = self.page_soup.find_all(class_="movimentacao")
        for movimentacao in self.movimentacoes:
            self.data = movimentacao.h1.text.replace("\n", "").replace("\t", "")
            movimentacao.h1.decompose()
            self.movimentacao = ' '.join(movimentacao.text.split())
            self.cur.execute("""INSERT INTO andamentos (numero_processo, data, andamento) VALUES 
                            (?, ?, ?)""", (processo, self.data, self.movimentacao))


class PesquisarProcessoPJE():
    '''Classe para realizar pesquisa no site PJE do TJMA dos processos listados em 'lista_processos'. 
    Os resultados serão salvos na pasta 'diretorio'. '''
    def __init__(self, lista_processos='', diretorio=''):
        self.lista_processos = lista_processos
        self.diretorio = diretorio
        self.driver = webdriver.Chrome('chromedriver.exe')
        self.driver.get('https://pje.tjma.jus.br/pje/ConsultaPublica/listView.seam')
        self.driver.implicitly_wait(3)
        if not os.path.exists(self.diretorio):
            os.makedirs("data/" + self.diretorio)
            os.makedirs("data/" + self.diretorio + "/html")
            os.makedirs("data/" + self.diretorio + "/error")
        self.open_db()
        try:
            self.contador = 1
            self.message = ""
            for processo, in self.lista_processos:
                while True:
                    self.inserido = self.inserir_numero(processo.value)
                    if self.message != "Resposta incorreta." and self.message != "Caracteres digitados de forma incorreta.":
                        break
                if "Consulta realizada com sucesso" in self.message:
                    self.driver.find_element_by_xpath("//img[@title='Ver Detalhes']").click()
                    self.driver.switch_to_window(self.driver.window_handles[-1])
                    self.html = self.driver.page_source
                    self.driver.close()
                    self.driver.switch_to_window(self.driver.window_handles[0])
                    with open("data/" + self.diretorio + '/html/' + str(processo.value) + '.html', 'w') as escrita:
                        escrita.write(self.html)
                    self.salva_banco(self.html, processo.value, self.message)
                self.contador = self.contador + 1
        except NoSuchElementException:
            if self.is_connected() == False:
                print("Sem conexão com a internet!")
            else:
                self.driver.save_screenshot("data/" + self.diretorio + "/error/last_exception.png")
                print_exc()
        except:
            self.driver.save_screenshot("data/" + self.diretorio + "/error/last_exception.png")
            print_exc()
        finally:
            self.conn.commit()
            self.conn.close()
            self.driver.close()
            self.driver.quit()
            pass

    def is_connected(self):
        try:
            create_connection(("www.google.com", 80))
            return True
        except OSError:
            pass
        return False

    def open_db(self):
        self.conn = sqlite3.connect("data/" + self.diretorio + '/' + self.diretorio + '.db')
        self.cur = self.conn.cursor()
        self.cur.execute('''CREATE TABLE IF NOT EXISTS processos
                    (id INTEGER PRIMARY KEY AUTOINCREMENT,
                    numero_processo TEXT UNIQUE, vara TEXT, dias_parados INTEGER, mensagem_download TEXT)''')
        self.cur.execute('''CREATE TABLE IF NOT EXISTS andamentos (numero_processo TEXT, data TEXT, 
                    andamento TEXT, FOREIGN KEY(numero_processo) REFERENCES processos(numero_processo))''')

    def inserir_numero(self, numero):
        self.campo_numero = self.driver.find_element_by_id('fPP:numProcessoDecoration:numProcesso')
        self.campo_numero.clear()
        self.campo_numero.send_keys(numero.replace("8.10",""))
        self.captcha_element = self.driver.find_element_by_id("fPP:j_id140:captchaImg")
        self.captcha = self.handle_captcha_pje(self.captcha_element, numero)
        self.driver.find_element_by_id('fPP:j_id140:verifyCaptcha').send_keys(self.captcha)
        self.driver.find_element_by_id('fPP:searchProcessos').click()
        try:
            #MENSAGEM DE QUE ERROU O CAPTCHA
            self.message = self.driver.find_element_by_xpath("//div[@id='fPP:j_id140:fieldDiv']/span[@class='errorFields errors']").text
        except:
            pass
        try:
            #MENSAGEM DE QUE NÃO ENCONTROU O PROCESSO
            self.message = self.driver.find_element_by_xpath("//*[contains(text(), 'Sua pesquisa não encontrou nenhum processo disponível.')]").text
        except:
            pass
        try:
            #MENSAGEM DE SUCESSO NA PESQUISA
            self.driver.find_element_by_xpath("//img[@title='Ver Detalhes']")
            self.message = "Consulta realizada com sucesso."
        except:
            pass
        print("Processo:", numero, "-", self.message)
        if self.message == "Resposta incorreta.":
            copyfile("data/" + self.diretorio+"/captcha_" + str(numero) + ".png", self.diretorio + "/error/captcha_" + str(numero) + ".png")
        os.remove("data/" + self.diretorio+"/captcha_" + str(numero) + ".png")

    def handle_captcha_pje(self, img, numero):
        self.screenshot_filename = "data/" + self.diretorio + "/screenshot_" + str(numero) + ".png"
        self.captcha_filename = "data/" + self.diretorio + "/captcha_" + str(numero) + ".png"
        self.location = img.location
        self.size = img.size
        self.driver.save_screenshot(self.screenshot_filename)
        self.screenshot = Image.open(self.screenshot_filename)
        self.left = self.location['x']
        self.top = self.location['y']
        self.right = self.location['x'] + self.size['width']
        self.bottom = self.location['y'] + self.size['height']
        self.captcha_img = self.screenshot.crop((self.left, self.top, self.right, self.bottom))
        self.screenshot.close()
        self.captcha_img, solution = solve_captcha_pje(self.captcha_img)
        self.captcha_img.save(self.captcha_filename)
        os.remove(self.screenshot_filename)
        return solution

    def salva_banco(self, html, processo, message):
        self.page_soup = BeautifulSoup(html, 'html.parser')
        self.vara = self.page_soup.find(text=re.compile('Vara')).strip()
        try:
            self.cur.execute("""INSERT INTO processos (numero_processo, vara, mensagem_download) VALUES
                            (?, ?, ?)""", (processo, self.vara, message))
        except sqlite3.IntegrityError:
            print_exc()
            return
        self.primeira_movimentacao = self.page_soup.find('span', {"id": re.compile('^j_id62:processoEvento:')}).text
        self.ultima_data = self.primeira_movimentacao[:10]
        self.qnt_dias = datetime.now() - datetime.strptime(self.ultima_data, "%d/%m/%Y")
        self.qnt_dias = str(self.qnt_dias.days)
        self.cur.execute("""UPDATE processos SET
                    dias_parados = ? WHERE numero_processo = ?""", (self.qnt_dias, processo))
        self.movimentacoes = self.page_soup.find_all('span', {"id": re.compile('^j_id62:processoEvento:')})
        for movimentacao in self.movimentacoes:
            self.data = movimentacao.text[:10]
            self.movimentacao = movimentacao.text.split(" - ")[1]
            self.cur.execute("""INSERT INTO andamentos (numero_processo, data, andamento) VALUES 
                            (?, ?, ?)""", (processo, self.data, self.movimentacao))


if __name__ == '__main__':
    wb = load_workbook(filename='excel/PROCESSOS - Pedreiras e Região.xlsx', read_only=True)
    ws = wb[wb.sheetnames[0]]
    #ws = wb['TODOS']
    range = ws['E3':'E960']
    IncluirPush = PesquisarProcessoJurisconsult(lista_processos=range,
                                            tipo_de_processo="primeiro_grau",
                                            diretorio='PROCESSOS - Pedreiras e Região')

    # wb = load_workbook(filename='excel/processos_descompressão_pje_mai_2017.xlsx', read_only=True)
    # ws = wb['PJE']
    # range = ws['C102':'C2199']
    # IncluirPush = PesquisarProcessoPJE(lista_processos=range,
    #                                     diretorio='processos_descompressão_pje_mai_2017')