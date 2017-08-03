import os
import traceback
from selenium import webdriver
from openpyxl import Workbook, load_workbook

class IncluirPushSelenium():
    '''Inclui processos indicados em 'lista_processos' no push do site jurisconsult,
    podendo ser processos de primeiro grau (1) ou segundo grau (2).'''
    def __init__(self, lista_processos='', usuario='', senha='', grau=1):
        self.lista_processos = lista_processos
        self.usuario = usuario
        self.senha = senha
        self.grau = grau
        self.driver = webdriver.Chrome('chromedriver.exe')
        self.driver.get('http://jurisconsult.tjma.jus.br')
        try:
            self.login()
            for processo, in self.lista_processos:
                #print(len(str(processo.value)), processo.value)
                if processo.value == None: continue
                self.inserir_processo(processo.value)
        finally:
            self.logout()
            self.driver.close()
            self.driver.quit()

    def login(self):
        self.driver.find_element_by_link_text('Push').click()
        self.driver.find_element_by_link_text('Login').click()
        self.driver.find_element_by_id('txtEmail').send_keys(self.usuario)
        self.driver.find_element_by_id('txtSenha').send_keys(self.senha)
        self.driver.find_element_by_id('btnEnviar').click()
        print("Logado no sistema PUSH!")

    def inserir_processo(self, numero):
        self.driver.find_element_by_link_text('Push').click()
        try:
            self.driver.find_element_by_link_text('Cadastrar Processo').click()
        except:
            traceback.print_exc()
            self.login()
            self.driver.find_element_by_link_text('Push').click()
            self.driver.find_element_by_link_text('Cadastrar Processo').click()
        if self.grau==1:
            #PROCESSOS DE PRIMEIRO GRAU
            self.driver.find_element_by_xpath('//select[@id="slcGrau"]/option[@value="1"]').click()
        elif self.grau==2:
            #PROCESSOS DE SEGUNDO GRAU
            self.driver.find_element_by_xpath('//select[@id="slcGrau"]/option[@value="2"]').click()
        if len(str(numero))<11:
            #NUMERAÇÃO CURTA
            self.driver.find_element_by_xpath('//select[@id="slcTipo"]/option[@value="1"]').click()
        else:
            #NUMERAÇÃO CNJ
            self.driver.find_element_by_xpath('//select[@id="slcTipo"]/option[@value="2"]').click()
        self.driver.find_element_by_id('txtChave').send_keys(numero)
        self.driver.find_element_by_id('Cadastrar').click()
        self.message = self.driver.find_element_by_class_name('message').text
        print(numero, self.message)
        return self.message

    def logout(self):
        self.driver.find_element_by_link_text('Push').click()
        self.driver.find_element_by_link_text('Cadastrar Processo').click()
        self.driver.find_element_by_link_text('SAIR').click()
        print("Logout realizado com sucesso no sistema PUSH!")

if __name__ == '__main__':
    wb = load_workbook(filename='excel/Precatórios 2017 03 ago 2017.xlsx', read_only=True)
    ws = wb['Plan1']
    range = ws['A1':'A25']
    IncluirPush = IncluirPushSelenium(lista_processos=range, 
        usuario=os.environ.get('JURISCONSULT_EMAIL', 'your-email@example.com'), 
        senha=os.environ.get('JURISCONSULT_PASSWORD', 'your-password'),
        grau=2)
